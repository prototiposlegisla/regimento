"""Fix all validation warnings in remissivo Teste.xlsx"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import re

import argparse
ap = argparse.ArgumentParser()
ap.add_argument("file", nargs="?", default="remissivo Teste.xlsx")
target = ap.parse_args().file
print(f"Fixing: {target}")
wb = openpyxl.load_workbook(target)
ws = wb["Sheet1"]

fixes_applied = 0

def fix_cell(row, col, old_val, new_val, description=""):
    global fixes_applied
    cell = ws.cell(row=row, column=col)
    if cell.value == old_val:
        cell.value = new_val
        fixes_applied += 1
        print(f"  Row {row}: {description or 'fixed'}")
        print(f"    OLD: {repr(old_val)}")
        print(f"    NEW: {repr(new_val)}")
    else:
        print(f"  Row {row}: SKIPPED - value doesn't match")
        print(f"    Expected: {repr(old_val)}")
        print(f"    Found:    {repr(cell.value)}")

# ============================================================
# 1. Global PU → §ú replacement in Dispositivos (column C=3)
# ============================================================
print("=== 1. PU → §ú replacement ===")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    if cell.value and isinstance(cell.value, str) and 'PU' in cell.value:
        old = cell.value
        # Replace PU with §ú - careful with word boundaries
        # PU can appear as: ,PU or start-of-line PU or PU, or PU end-of-line
        new = re.sub(r'\bPU\b', '§ú', old)
        if new != old:
            cell.value = new
            fixes_applied += 1
            print(f"  Row {row}: PU → §ú")

# ============================================================
# 2. Art.47 missing alínea 'a': 47,ROMAN,NUMBER → 47,ROMAN,a,NUMBER
# ============================================================
print("\n=== 2. Art.47 missing alínea 'a' ===")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    if cell.value and isinstance(cell.value, str) and '47,' in cell.value:
        old = cell.value
        lines = old.split('\n')
        new_lines = []
        changed = False
        for line in lines:
            # Match 47,ROMAN,NUMBER where NUMBER is a plain digit (item)
            # e.g. 47,V,4 → 47,V,a,4
            new_line = re.sub(r'^(47,[IVXLC]+),(\d+)$', r'\1,a,\2', line)
            if new_line != line:
                changed = True
            new_lines.append(new_line)
        if changed:
            new = '\n'.join(new_lines)
            cell.value = new
            fixes_applied += 1
            print(f"  Row {row}: added alínea 'a'")
            print(f"    OLD: {repr(old)}")
            print(f"    NEW: {repr(new)}")

# ============================================================
# 3. ROMAN,§ splitting: ART,ROMAN,§N → ART,ROMAN\nART,§N
#    But NOT §ú,ROMAN (that's legitimate nesting)
# ============================================================
print("\n=== 3. ROMAN,§ splitting ===")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    if cell.value and isinstance(cell.value, str):
        old = cell.value
        lines = old.split('\n')
        new_lines = []
        changed = False
        for line in lines:
            # Match pattern: ART,ROMAN,§N (inciso followed by paragraph)
            # e.g. 34,II,§1 → 34,II\n34,§1
            # But NOT: ART,§ú,ROMAN (paragraph with nested inciso - legitimate)
            m = re.match(r'^(\d+(?:-[A-Z])?),([IVXLC]+),(§.+)$', line)
            if m:
                art = m.group(1)
                roman = m.group(2)
                para = m.group(3)
                new_lines.append(f"{art},{roman}")
                new_lines.append(f"{art},{para}")
                changed = True
            else:
                new_lines.append(line)
        if changed:
            new = '\n'.join(new_lines)
            cell.value = new
            fixes_applied += 1
            print(f"  Row {row}: split ROMAN,§")
            print(f"    OLD: {repr(old)}")
            print(f"    NEW: {repr(new)}")

# ============================================================
# 4. Specific broken references
# ============================================================
print("\n=== 4. Specific broken references ===")

# Row 231: '357,358,359' → '357\n358\n359'
fix_cell(231, 3, '357,358,359', '357\n358\n359', "split concatenated articles")

# Row 310: '1-6DT' → '1-6' (ADT articles 1 through 6, remove DT suffix)
fix_cell(310, 3, '1-6DT', '1-6', "remove DT suffix from ADT range")

# Row 369: '233,§2\n234,V e §2\n315-323' → '233,§2\n234,V\n234,§2\n315-323'
fix_cell(369, 3, '233,§2\n234,V e §2\n315-323', '233,§2\n234,V\n234,§2\n315-323',
         "split 'V e §2'")

# Row 377: '370,I e 371' → '370,I\n371'
fix_cell(377, 3, '370,I e 371', '370,I\n371', "split 'I e 371'")

# Row 424: '4,5,7\n9\n105,I\n289' → '4\n5\n7\n9\n105,I\n289'
fix_cell(424, 3, '4,5,7\n9\n105,I\n289', '4\n5\n7\n9\n105,I\n289',
         "split concatenated articles 4,5,7")

# Row 552: '105,IV\nV\nVII' → '105,IV\n105,V\n105,VII'
fix_cell(552, 3, '105,IV\nV\nVII', '105,IV\n105,V\n105,VII',
         "expand orphan Roman numerals")

# Row 638: '171,§4\n239,§1,248' → '171,§4\n239,§1\n248'
fix_cell(638, 3, '171,§4\n239,§1,248', '171,§4\n239,§1\n248',
         "split concatenated §1,248")

# Row 852: '156\n157§5\n160,§6' → '156\n157,§5\n160,§6'
fix_cell(852, 3, '156\n157§5\n160,§6', '156\n157,§5\n160,§6',
         "add comma in 157§5")

# Row 891: '183-A,I a IV,§2' → '183-A,I\n183-A,II\n183-A,III\n183-A,IV\n183-A,§2'
fix_cell(891, 3, '183-A,I a IV,§2',
         '183-A,I\n183-A,II\n183-A,III\n183-A,IV\n183-A,§2',
         "expand 'I a IV' and split §2")

# Row 943: '4-D-4-F' → '4-D\n4-E\n4-F'
# These are ADT articles 4-D, 4-E, 4-F
fix_cell(943, 3, '4-D-4-F', '4-D\n4-E\n4-F', "expand ADT article range")

# Row 950: '211,VIII\n269a273' → '211,VIII\n269-273'
fix_cell(950, 3, '211,VIII\n269a273', '211,VIII\n269-273',
         "fix range notation 269a273")

# Row 966: '217,§1,2e3' → '217,§1\n217,§2\n217,§3'
fix_cell(966, 3, '217,§1,2e3', '217,§1\n217,§2\n217,§3',
         "expand §1,2e3 to individual paragraphs")

# Row 1065: '140,XI a XIII' → '140,XI\n140,XII\n140,XIII'
fix_cell(1065, 3, '140,XI a XIII', '140,XI\n140,XII\n140,XIII',
         "expand 'XI a XIII'")

# ============================================================
# 5. Scan for any remaining orphan Roman numerals (V, VII etc alone on a line)
# ============================================================
print("\n=== 5. Scan for remaining orphan lines ===")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    if cell.value and isinstance(cell.value, str):
        lines = cell.value.split('\n')
        if len(lines) > 1:
            for i, line in enumerate(lines):
                # Check for orphan Roman numerals (no article prefix)
                if re.match(r'^[IVXLC]+$', line.strip()):
                    # Find the article number from the previous line
                    for j in range(i - 1, -1, -1):
                        m = re.match(r'^(\d+(?:-[A-Z])?)', lines[j])
                        if m:
                            art = m.group(1)
                            old_val = cell.value
                            lines[i] = f"{art},{line.strip()}"
                            cell.value = '\n'.join(lines)
                            fixes_applied += 1
                            print(f"  Row {row}: orphan '{line.strip()}' → '{art},{line.strip()}'")
                            break

# ============================================================
# 6. Scan for remaining 'e' connectors in references
# ============================================================
print("\n=== 6. Scan for remaining 'e' connectors ===")
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)
    if cell.value and isinstance(cell.value, str):
        lines = cell.value.split('\n')
        new_lines = []
        changed = False
        for line in lines:
            # Pattern: ART,ROMAN e ROMAN → split
            m = re.match(r'^(\d+(?:-[A-Z])?),([IVXLC]+)\s+e\s+([IVXLC]+)$', line)
            if m:
                art, r1, r2 = m.group(1), m.group(2), m.group(3)
                new_lines.append(f"{art},{r1}")
                new_lines.append(f"{art},{r2}")
                changed = True
                continue
            # Pattern: ART,§N e §M → split
            m = re.match(r'^(\d+(?:-[A-Z])?),(§\S+)\s+e\s+(§\S+)$', line)
            if m:
                art, p1, p2 = m.group(1), m.group(2), m.group(3)
                new_lines.append(f"{art},{p1}")
                new_lines.append(f"{art},{p2}")
                changed = True
                continue
            new_lines.append(line)
        if changed:
            old = cell.value
            cell.value = '\n'.join(new_lines)
            fixes_applied += 1
            print(f"  Row {row}: split 'e' connector")
            print(f"    OLD: {repr(old)}")
            print(f"    NEW: {repr(cell.value)}")

# ============================================================
# Save
# ============================================================
print(f"\n=== Total fixes applied: {fixes_applied} ===")
wb.save(target)
print(f"Saved {target}")
