#!/usr/bin/env python3
"""Merge LOM remissive index from PDF into existing remissivo.xlsx.

Reads the Biblioteca's PDF index for the Lei Orgânica, parses every entry,
converts article references to spreadsheet notation (LOM:art,detail), and
merges into the existing remissivo.xlsx — adding new subjects and
complementing existing ones with LOM articles.
"""

import re
import sys
import shutil
from pathlib import Path
from collections import defaultdict

import pdfplumber
import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────
PDF_PATH = Path(r"C:\Users\kauen\Documents\ferramentas\legislasampa\regimento\remissivo\índice remissibo LOM.pdf")
XLSX_PATH = Path(r"C:\Users\kauen\OneDrive\CMSP\Regimento\Fontes\remissivo.xlsx")
BACKUP_PATH = XLSX_PATH.with_suffix(".xlsx.bak")

ROMAN = re.compile(r'^[IVXLC]+$')
ROMAN_LIST = ['I','II','III','IV','V','VI','VII','VIII','IX','X',
              'XI','XII','XIII','XIV','XV','XVI','XVII','XVIII','XIX','XX',
              'XXI','XXII','XXIII','XXIV','XXV','XXVI','XXVII','XXVIII','XXIX','XXX']

def roman_to_int(r):
    vals = {'I':1,'V':5,'X':10,'L':50,'C':100}
    total = 0
    for i, c in enumerate(r):
        if i+1 < len(r) and vals[c] < vals[r[i+1]]:
            total -= vals[c]
        else:
            total += vals[c]
    return total

def int_to_roman(n):
    if 1 <= n <= len(ROMAN_LIST):
        return ROMAN_LIST[n-1]
    # fallback
    result = ''
    for val, sym in [(100,'C'),(90,'XC'),(50,'L'),(40,'XL'),(10,'X'),(9,'IX'),(5,'V'),(4,'IV'),(1,'I')]:
        while n >= val:
            result += sym
            n -= val
    return result


# ═════════════════════════════════════════════════════════════════════════
# Step 1: Extract text from PDF (layout-aware)
# ═════════════════════════════════════════════════════════════════════════

# Indent levels (leading spaces from layout=True extraction):
#   21+  → page header / page number  (skip)
#   19-20 → ver-target continuation (level 2)
#   13-15 → content: sub-topic, article, ver (level 1)
#   10-12 → main entry or cross-ref heading (level 0)

LEVEL_ENTRY = 0    # main assunto / cross-ref heading
LEVEL_CONTENT = 1  # sub-topic, article, ver
LEVEL_VER_TGT = 2  # "ver também" target (deep indent)

def _classify_indent(leading_spaces: int) -> int:
    """Map leading-space count to semantic level."""
    if leading_spaces >= 19:
        return LEVEL_VER_TGT
    if leading_spaces >= 12:
        return LEVEL_CONTENT
    return LEVEL_ENTRY


def _is_skip_line(text: str) -> bool:
    """Lines to skip entirely (headers, footers, section letters)."""
    if 'LEI ORGÂNICA DO MUNICÍPIO DE SÃO PAULO' in text:
        return True
    if re.match(r'^\d+$', text):
        return True
    if 'NDICE' in text and 'REMISSIVO' in text:
        return True
    if re.match(r'^_+$', text):
        return True
    if re.match(r'^[A-Z]$', text):
        return True
    return False


def _is_all_caps(text: str) -> bool:
    """Check if text is predominantly uppercase."""
    letters = [c for c in text if c.isalpha()]
    if not letters:
        return False
    return sum(1 for c in letters if c.isupper()) / len(letters) > 0.8


def extract_lines(pdf_path):
    """Extract lines from PDF using layout-preserving mode.

    Returns list of (level, text) tuples where level is
    LEVEL_ENTRY, LEVEL_CONTENT, or LEVEL_VER_TGT.
    Continuation lines are already joined.
    """
    # First pass: collect raw (leading_spaces, stripped_text) for all pages
    raw_items = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(layout=True)
            if not text:
                continue
            for line in text.split('\n'):
                rline = line.rstrip()
                stripped = rline.strip()
                if not stripped:
                    continue
                if _is_skip_line(stripped):
                    continue
                leading = len(rline) - len(rline.lstrip())
                level = _classify_indent(leading)
                raw_items.append((level, stripped))

    # Second pass: join continuation lines.
    # A continuation is an ENTRY-level line that should be appended to the
    # previous line rather than starting a new entry.  Two patterns:
    #   (a) Starts with lowercase → always a continuation of the previous line
    #   (b) ALL CAPS at entry level, but previous entry-level line ended with
    #       a conjunction/preposition word (E, OU, DE, DO, DA, etc.)
    CONJUNCTIONS = {'E', 'OU', 'DE', 'DO', 'DA', 'DOS', 'DAS', 'EM', 'A', 'O', 'AO', 'À'}

    joined = []  # list of (level, text)
    for level, text in raw_items:
        if level == LEVEL_ENTRY and joined:
            # Check if this is a continuation
            first_char = text[0] if text else ''
            is_continuation = False

            if first_char.islower():
                # Pattern (a): lowercase start → continuation of previous line
                is_continuation = True
            elif not _is_all_caps(text) and joined:
                # Pattern (c): capitalized word (e.g., "Municipal") continuing
                # a sub-topic line (starts with "-") at content level.
                # Example: "- remuneração: Procurador do Instituto de Previdência"
                #          "Municipal"  ← this is a continuation, not a new entry
                prev_level, prev_text = joined[-1]
                if prev_level == LEVEL_CONTENT and prev_text.startswith('- '):
                    is_continuation = True
            elif _is_all_caps(text):
                # Pattern (b): previous entry-level ALL CAPS ending with conjunction
                # Walk back to find the last entry-level line
                for j in range(len(joined) - 1, -1, -1):
                    if joined[j][0] == LEVEL_ENTRY:
                        prev_text = joined[j][1]
                        last_word = prev_text.rstrip().rsplit(None, 1)[-1] if prev_text.strip() else ''
                        if last_word.upper() in CONJUNCTIONS:
                            is_continuation = True
                        break

            if is_continuation:
                # Find the line to append to (walk back, could be any level)
                if joined:
                    prev_level, prev_text = joined[-1]
                    joined[-1] = (prev_level, prev_text + ' ' + text)
                    continue

        joined.append((level, text))

    # Third pass: fix L0 lines sandwiched between L2 lines (ver-target wrap).
    # Example: "FUNDAÇÃO INSTITUÍDA OU MANTIDA PELO PODER" (L2) →
    #          "MUNICIPAL" (L0) → "SOCIEDADE DE ECONOMIA MISTA" (L2)
    # The L0 line is actually a continuation of the previous L2 line.
    fixed = []
    for i, (level, text) in enumerate(joined):
        if level == LEVEL_ENTRY and _is_all_caps(text):
            prev_level = fixed[-1][0] if fixed else -1
            next_level = joined[i + 1][0] if i + 1 < len(joined) else -1
            if prev_level == LEVEL_VER_TGT and next_level == LEVEL_VER_TGT:
                prev_l, prev_t = fixed[-1]
                fixed[-1] = (prev_l, prev_t + ' ' + text)
                continue
        fixed.append((level, text))

    return fixed


# ═════════════════════════════════════════════════════════════════════════
# Step 2: Parse text into structured entries
# ═════════════════════════════════════════════════════════════════════════

def is_art_line(text):
    """Check if text is an article reference."""
    return bool(re.match(r'^art[\.\,]\s*\d', text, re.IGNORECASE))


def is_sub_topic(text):
    """Check if text starts with '- ' indicating a sub-topic."""
    return text.startswith('- ') or (text.startswith('-') and len(text) > 1 and text[1] == ' ')


def is_ver_line(text):
    """Check if text is a 'ver' or 'ver também' cross-reference."""
    low = text.lower()
    return low.startswith('ver ') or low.startswith('ver\t')


def parse_entries(items):
    """Parse (level, text) items into structured entries.

    Returns list of dicts:
    {
        'assunto': str,
        'is_cross_ref_only': bool,
        'sub_entries': [
            {
                'sub_assunto': str,
                'articles': [str],
                'vides': [str],
            }
        ]
    }
    """
    entries = []
    current_entry = None
    current_sub = None
    in_ver_block = False
    ver_targets = []

    # ver_base_target: the main target from "ver TARGET" line (e.g., "INDÚSTRIA")
    ver_base_target = ''

    def flush_ver_block():
        nonlocal in_ver_block, ver_targets, ver_base_target
        if in_ver_block and ver_targets and current_sub is not None:
            current_sub['vides'].extend(ver_targets)
        in_ver_block = False
        ver_targets = []
        ver_base_target = ''

    def new_sub(sub_text=''):
        nonlocal current_sub
        flush_ver_block()
        current_sub = {'sub_assunto': sub_text, 'articles': [], 'vides': []}
        if current_entry:
            current_entry['sub_entries'].append(current_sub)
        return current_sub

    for level, text in items:
        if not text:
            continue

        # ── Level 2: ver-target continuation ──
        if level == LEVEL_VER_TGT:
            if is_sub_topic(text):
                # Sub-detail of ver target: "- vantagens"
                # Creates a NEW vide "TARGET|detail" for the last known target
                detail = text.lstrip('- ').strip()
                if in_ver_block and ver_targets:
                    last_target = ver_targets[-1]
                    # If last target already has a pipe, it already got a sub-detail;
                    # this new sub-detail creates a separate vide
                    base = last_target.split('|')[0]
                    ver_targets.append(f'{base}|{detail}')
                    # Remove the bare target if it now has a sub-detail
                    if last_target == base:
                        ver_targets.remove(base)
                elif current_sub and current_sub['vides']:
                    last_vide = current_sub['vides'][-1]
                    base = last_vide.split('|')[0]
                    current_sub['vides'].append(f'{base}|{detail}')
                    if last_vide == base:
                        current_sub['vides'].remove(base)
            else:
                # ALL CAPS target after "ver SOMETHING"
                if in_ver_block:
                    ver_targets.append(text)
                elif current_sub:
                    current_sub['vides'].append(text)
            continue

        # ── Level 1: content (sub-topic, article, ver) ──
        if level == LEVEL_CONTENT:
            # "ver" / "ver também"
            if is_ver_line(text):
                flush_ver_block()
                m = re.match(r'^ver\s+também\s+(.+)$', text, re.IGNORECASE)
                if m:
                    target = m.group(1).strip()
                    if current_sub is None and current_entry:
                        current_sub = new_sub('')
                    if current_sub:
                        current_sub['vides'].append(target)
                        in_ver_block = True
                        ver_targets = []
                        ver_base_target = target
                    continue
                m = re.match(r'^ver\s+(.+)$', text, re.IGNORECASE)
                if m:
                    target = m.group(1).strip()
                    if current_sub is None and current_entry:
                        current_sub = new_sub('')
                    if current_sub:
                        current_sub['vides'].append(target)
                        in_ver_block = True
                        ver_targets = []
                        ver_base_target = target
                    continue

            # Article reference
            if is_art_line(text):
                flush_ver_block()
                if current_sub is None and current_entry:
                    current_sub = new_sub('')
                if current_sub:
                    current_sub['articles'].append(text)
                continue

            # Sub-topic
            if is_sub_topic(text):
                sub_text = text.lstrip('- ').strip()

                # If we're in a ver block, create a new vide "TARGET|sub_text"
                if in_ver_block:
                    if ver_targets:
                        base = ver_targets[-1].split('|')[0]
                        ver_targets.append(f'{base}|{sub_text}')
                        # Remove bare target if it now has a sub-detail
                        bare = base
                        if bare in ver_targets:
                            ver_targets.remove(bare)
                    elif current_sub and current_sub['vides']:
                        base = current_sub['vides'][-1].split('|')[0]
                        current_sub['vides'].append(f'{base}|{sub_text}')
                        if base in current_sub['vides']:
                            current_sub['vides'].remove(base)
                    continue

                # If sub-topic text is actually "art. N", treat as article ref
                if is_art_line(sub_text):
                    if current_sub is None and current_entry:
                        current_sub = new_sub('')
                    if current_sub:
                        current_sub['articles'].append(sub_text)
                    continue

                flush_ver_block()
                if current_entry:
                    current_sub = new_sub(sub_text)
                continue

            # ALL CAPS text at content level during ver block → additional ver target
            if in_ver_block and _is_all_caps(text):
                ver_targets.append(text)
                continue

            # Fallback for content-level lines that don't match patterns
            if current_sub and re.match(r'^art', text, re.IGNORECASE):
                current_sub['articles'].append(text)
            continue

        # ── Level 0: main entry or cross-ref heading ──
        flush_ver_block()

        if _is_all_caps(text) and not text.lower().startswith(('art.', 'art,', 'ver ')):
            # ALL CAPS → main assunto
            current_entry = {
                'assunto': text,
                'is_cross_ref_only': False,
                'sub_entries': []
            }
            entries.append(current_entry)
            current_sub = None
        else:
            # Mixed case → cross-reference heading (e.g., "Cargo em Comissão")
            current_entry = {
                'assunto': text,
                'is_cross_ref_only': True,
                'sub_entries': []
            }
            entries.append(current_entry)
            current_sub = new_sub('')

    flush_ver_block()
    return entries


# ═════════════════════════════════════════════════════════════════════════
# Step 3: Convert article references to spreadsheet format
# ═════════════════════════════════════════════════════════════════════════

def parse_art_reference(raw):
    """Parse a raw article reference from the PDF into spreadsheet notation.

    Returns list of strings like ["LOM:80", "LOM:80,I", "LOM:80,§1"].
    """
    raw = raw.strip()

    # Remove leading "art." or "art,"
    raw = re.sub(r'^art[\.\,]\s*', '', raw, flags=re.IGNORECASE)

    # Detect and remove D.G.T. / D.T. / D.T suffix → ADT prefix
    is_adt = False
    if re.search(r'\bD\.?G\.?T\.?\s*$', raw):
        is_adt = True
        raw = re.sub(r'\s*D\.?G\.?T\.?\s*$', '', raw)
    elif re.search(r'\bD\.?T\.?\s*$', raw):
        is_adt = True
        raw = re.sub(r'\s*D\.?T\.?\s*$', '', raw)

    # Normalize stray spaces before ordinal markers: "6 º" → "6º"
    raw = re.sub(r'(\d)\s+([°ºª])', r'\1\2', raw)

    # Extract article number
    # Pattern: digits + optional º/ª + optional -LETTER
    m = re.match(r'^(\d+)\s*[°ºª]?\s*(-[A-Z])?\s*(.*)$', raw)
    if not m:
        return []

    art_num = m.group(1)
    art_letter = m.group(2) or ''  # e.g., "-A"
    remainder = m.group(3).strip()

    art_id = art_num + art_letter  # e.g., "80" or "55-A"
    if is_adt:
        art_id = 'ADT' + art_id

    prefix = 'LOM:'

    # If no remainder, just the article
    if not remainder:
        return [f'{prefix}{art_id}']

    # Remove leading comma
    remainder = remainder.lstrip(',').strip()

    if not remainder:
        return [f'{prefix}{art_id}']

    # Parse the detail part
    return _parse_details(prefix, art_id, remainder)


def _parse_details(prefix, art_id, detail_str):
    """Parse detail string into individual references."""
    refs = []

    # Normalize stray spaces before ordinal markers
    detail_str = re.sub(r'(\d)\s+([°ºª])', r'\1\2', detail_str)

    # Handle "parágrafo único" / "§ único"
    if 'parágrafo único' in detail_str.lower() or '§ único' in detail_str.lower():
        detail_str_clean = re.sub(r',?\s*(?:parágrafo único|§\s*único)', '', detail_str, flags=re.IGNORECASE).strip()
        detail_str_clean = detail_str_clean.strip(',').strip()
        refs.append(f'{prefix}{art_id},§ú')
        if detail_str_clean:
            refs.extend(_parse_details(prefix, art_id, detail_str_clean))
        return refs

    # Handle "caput" mixed with other details
    if 'caput' in detail_str.lower():
        detail_str_clean = re.sub(r',?\s*caput', '', detail_str, flags=re.IGNORECASE).strip()
        detail_str_clean = detail_str_clean.strip(',').strip()
        refs.append(f'{prefix}{art_id},caput')
        if detail_str_clean:
            refs.extend(_parse_details(prefix, art_id, detail_str_clean))
        return refs

    # Handle "§ N a § M" paragraph ranges: "§ 1º a § 4º" → §1, §2, §3, §4
    range_para_m = re.match(r'^§\s*(\d+)[°ºª]?\s*a\s*§\s*(\d+)[°ºª]?\s*(.*)$', detail_str)
    if range_para_m:
        start_p = int(range_para_m.group(1))
        end_p = int(range_para_m.group(2))
        rest = range_para_m.group(3).strip().lstrip(',').strip()
        for n in range(start_p, end_p + 1):
            refs.append(f'{prefix}{art_id},§{n}')
        if rest:
            refs.extend(_parse_details(prefix, art_id, rest))
        return refs

    # Handle "§§ Nº, Mº" (multiple paragraphs listed)
    m = re.match(r'^§§\s*(.+)$', detail_str)
    if m:
        nums_str = m.group(1)
        for pm in re.finditer(r'(\d+)[°ºª]?', nums_str):
            refs.append(f'{prefix}{art_id},§{pm.group(1)}')
        return refs if refs else [f'{prefix}{art_id}']

    # Handle single "§ Nº" possibly followed by more details
    m = re.match(r'^§\s*(\d+)[°ºª]?\s*(.*)$', detail_str)
    if m:
        para_num = m.group(1)
        rest = m.group(2).strip().lstrip(',').strip()
        if rest:
            # Check if remainder contains "§§" (additional paragraphs at article level)
            # e.g., "I, II, §§ 7º, 8º" → §5,I and §5,II are nested, §7 and §8 are siblings
            para_break = re.search(r'(?:,\s*)?§§\s*(.+)$', rest)
            if para_break:
                nested_part = rest[:para_break.start()].strip().rstrip(',').strip()
                sibling_part = para_break.group(1).strip()
                # Process nested sub-details of this paragraph
                if nested_part:
                    sub_details = _split_details(nested_part)
                    for sd in sub_details:
                        refs.append(f'{prefix}{art_id},§{para_num},{sd}')
                else:
                    refs.append(f'{prefix}{art_id},§{para_num}')
                # Process sibling paragraphs
                for pm in re.finditer(r'(\d+)[°ºª]?', sibling_part):
                    refs.append(f'{prefix}{art_id},§{pm.group(1)}')
            else:
                sub_details = _split_details(rest)
                for sd in sub_details:
                    # If the sub-detail is a § reference, it's a sibling paragraph
                    sd_para = re.match(r'^§(\d+)$', sd)
                    if sd_para:
                        refs.append(f'{prefix}{art_id},§{sd_para.group(1)}')
                    else:
                        refs.append(f'{prefix}{art_id},§{para_num},{sd}')
        else:
            refs.append(f'{prefix}{art_id},§{para_num}')
        return refs

    # Handle "I a IV" ranges of roman numerals
    range_m = re.match(r'^([IVXLC]+)\s+a\s+([IVXLC]+)\s*(.*)$', detail_str)
    if range_m:
        start_r = roman_to_int(range_m.group(1))
        end_r = roman_to_int(range_m.group(2))
        rest = range_m.group(3).strip().lstrip(',').strip()
        for n in range(start_r, end_r + 1):
            refs.append(f'{prefix}{art_id},{int_to_roman(n)}')
        if rest:
            refs.extend(_parse_details(prefix, art_id, rest))
        return refs

    # Split by comma and handle each part
    parts = _split_details(detail_str)

    if len(parts) == 1:
        refs.append(f'{prefix}{art_id},{parts[0]}')
    elif all(ROMAN.match(p) for p in parts):
        for p in parts:
            refs.append(f'{prefix}{art_id},{p}')
    elif len(parts) >= 2:
        types = []
        for p in parts:
            if ROMAN.match(p):
                types.append('roman')
            elif re.match(r'^[a-z]$', p):
                types.append('letter')
            elif re.match(r'^\d+$', p) and int(p) < 20:
                types.append('number')
            elif re.match(r'^§\d+$', p):
                types.append('para')
            else:
                types.append('other')

        hierarchy = {'roman': 0, 'letter': 1, 'number': 2, 'para': -1, 'other': -1}

        i = 0
        while i < len(parts):
            # Paragraph references are always siblings at article level
            if types[i] == 'para':
                refs.append(f'{prefix}{art_id},{parts[i]}')
                i += 1
                continue
            nested = [parts[i]]
            j = i + 1
            while j < len(parts):
                if types[j] == 'para':
                    break  # paragraphs break nesting
                prev_h = hierarchy.get(types[j-1], -1)
                curr_h = hierarchy.get(types[j], -1)
                if curr_h > prev_h and prev_h >= 0 and curr_h >= 0:
                    nested.append(parts[j])
                    j += 1
                else:
                    break
            refs.append(f'{prefix}{art_id},{",".join(nested)}')
            i = j

    return refs if refs else [f'{prefix}{art_id}']


def _split_details(detail_str):
    """Split a detail string into individual components.

    Handles comma separation and "e" connector.
    """
    # First handle "e" connector: "III e V" → "III, V"
    detail_str = re.sub(r'\s+e\s+', ', ', detail_str)

    # Split by comma
    parts = [p.strip() for p in detail_str.split(',') if p.strip()]

    # Clean up each part
    cleaned = []
    for p in parts:
        # Remove ordinal markers
        p = re.sub(r'[°ºª]', '', p).strip()
        # Normalize "§ N" → "§N" (remove space between § and digit)
        p = re.sub(r'^§\s+(\d)', r'§\1', p)
        if p:
            cleaned.append(p)

    return cleaned


# ═════════════════════════════════════════════════════════════════════════
# Step 4: Convert parsed entries to spreadsheet rows
# ═════════════════════════════════════════════════════════════════════════

def entries_to_rows(entries):
    """Convert parsed entries to spreadsheet rows.

    Returns list of (assunto, sub_assunto, dispositivos_str, vides_str).
    """
    rows = []

    for entry in entries:
        assunto = entry['assunto']

        if entry['is_cross_ref_only']:
            # Cross-reference entry: just vides, no articles
            for sub in entry['sub_entries']:
                vides_str = '\n'.join(sub['vides']) if sub['vides'] else ''
                if vides_str:
                    rows.append((assunto, '', '', vides_str))
            if not entry['sub_entries']:
                rows.append((assunto, '', '', ''))
            continue

        if not entry['sub_entries']:
            # Entry with no sub-entries at all (shouldn't happen normally)
            rows.append((assunto, '', '', ''))
            continue

        for sub in entry['sub_entries']:
            sub_assunto = sub['sub_assunto']

            # Convert articles
            all_refs = []
            for art_raw in sub['articles']:
                refs = parse_art_reference(art_raw)
                all_refs.extend(refs)

            dispositivos_str = '\n'.join(all_refs) if all_refs else ''
            vides_str = '\n'.join(sub['vides']) if sub['vides'] else ''

            rows.append((assunto, sub_assunto, dispositivos_str, vides_str))

    return rows


# ═════════════════════════════════════════════════════════════════════════
# Step 5: Merge with existing spreadsheet
# ═════════════════════════════════════════════════════════════════════════

def normalize_key(s):
    """Normalize a string for comparison (lowercase, strip, collapse whitespace)."""
    return re.sub(r'\s+', ' ', s.strip().lower())


def merge_into_xlsx(xlsx_path, new_rows, output_path):
    """Merge new LOM rows into existing spreadsheet.

    For existing assunto+sub_assunto: append LOM articles to Dispositivos.
    For new assunto or new sub_assunto: insert new rows.
    """
    wb = openpyxl.load_workbook(str(xlsx_path))

    # Find the main sheet (first non-Normas)
    ws = None
    for name in wb.sheetnames:
        if name != "Normas":
            ws = wb[name]
            break

    if ws is None:
        print("ERROR: No main sheet found!")
        return

    # Build index of existing entries: (normalized_assunto, normalized_sub) → row_number
    existing = {}  # (norm_assunto, norm_sub) → row_num
    existing_assuntos = set()  # just the assuntos

    for row_num in range(2, ws.max_row + 1):
        assunto = str(ws.cell(row_num, 1).value or '').strip()
        sub = str(ws.cell(row_num, 2).value or '').strip()
        if assunto:
            key = (normalize_key(assunto), normalize_key(sub))
            existing[key] = row_num
            existing_assuntos.add(normalize_key(assunto))

    # Track what we'll append and what we'll insert
    appended = 0
    new_subs_added = 0
    new_assuntos_added = 0
    rows_to_insert = []  # (assunto, sub, dispositivos, vides) for truly new entries

    for (assunto, sub, dispositivos, vides) in new_rows:
        if not assunto:
            continue

        norm_assunto = normalize_key(assunto)
        norm_sub = normalize_key(sub)
        key = (norm_assunto, norm_sub)

        if key in existing:
            # Existing entry: append LOM articles
            row_num = existing[key]
            if dispositivos:
                current_disp = str(ws.cell(row_num, 3).value or '').strip()
                if current_disp:
                    # Check for duplicates before appending
                    existing_refs = set(r.strip() for r in current_disp.split('\n') if r.strip())
                    new_refs = [r.strip() for r in dispositivos.split('\n') if r.strip()]
                    added_refs = [r for r in new_refs if r not in existing_refs]
                    if added_refs:
                        ws.cell(row_num, 3).value = current_disp + '\n' + '\n'.join(added_refs)
                        appended += 1
                else:
                    ws.cell(row_num, 3).value = dispositivos
                    appended += 1

            if vides:
                current_vides = str(ws.cell(row_num, 4).value or '').strip()
                if current_vides:
                    existing_vides = set(v.strip() for v in current_vides.split('\n') if v.strip())
                    new_vides = [v.strip() for v in vides.split('\n') if v.strip()]
                    added_vides = [v for v in new_vides if v not in existing_vides]
                    if added_vides:
                        ws.cell(row_num, 4).value = current_vides + '\n' + '\n'.join(added_vides)
                else:
                    ws.cell(row_num, 4).value = vides

        elif norm_assunto in existing_assuntos:
            # Assunto exists but this specific sub_assunto is new
            # Find where to insert (after the last row of this assunto)
            rows_to_insert.append((assunto, sub, dispositivos, vides, 'new_sub'))
            new_subs_added += 1

        else:
            # Completely new assunto
            rows_to_insert.append((assunto, sub, dispositivos, vides, 'new_assunto'))
            new_assuntos_added += 1

    # Now insert new rows
    # First, find insertion points for new sub-assuntos within existing assuntos
    # Then find insertion points for new assuntos (alphabetical order)

    if rows_to_insert:
        # Group new rows by assunto for insertion
        # We'll append all new rows at the end and let the user sort if needed
        # (the instructions say "A ordenação final das entradas no índice é feita
        #  alfabeticamente pelo sistema, então a ordem das linhas na planilha não
        #  precisa ser alfabética.")

        next_row = ws.max_row + 1
        for item in rows_to_insert:
            assunto, sub, dispositivos, vides = item[0], item[1], item[2], item[3]
            ws.cell(next_row, 1).value = assunto
            ws.cell(next_row, 2).value = sub if sub else None
            ws.cell(next_row, 3).value = dispositivos if dispositivos else None
            ws.cell(next_row, 4).value = vides if vides else None
            next_row += 1

    # Save
    wb.save(str(output_path))
    wb.close()

    return appended, new_subs_added, new_assuntos_added


# ═════════════════════════════════════════════════════════════════════════
# Step 6: Handle article ranges in PDF ("art. 32 a 35 D.G.T.")
# ═════════════════════════════════════════════════════════════════════════

def expand_art_range_line(line):
    """Expand article range references into individual article lines.

    "art. 32 a 35 D.G.T." → ["art. 32 D.G.T.", "art. 33 D.G.T.", ...]
    "art. 32 a 35 D.T" → ["art. 32 D.T", "art. 33 D.T", ...]
    """
    stripped = line.strip()

    # Pattern: art. N a M [D.G.T.|D.T]
    m = re.match(
        r'^(art[\.\,])\s*(\d+)\s*[°ºª]?\s+a\s+(\d+)\s*[°ºª]?\s*(D\.?G?\.?T\.?)?\s*$',
        stripped, re.IGNORECASE
    )
    if m:
        prefix = m.group(1)
        start = int(m.group(2))
        end = int(m.group(3))
        suffix = m.group(4) or ''
        result = []
        for n in range(start, end + 1):
            result.append(f'{prefix} {n} {suffix}'.strip())
        return result

    return [stripped]


# ═════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("Merge LOM Remissive Index into remissivo.xlsx")
    print("=" * 60)

    # Step 1: Extract text (layout-aware)
    print(f"\n[1/5] Extracting text from PDF (layout-aware)...")
    items = extract_lines(PDF_PATH)
    print(f"  Extracted {len(items)} items (level, text)")

    # Expand article ranges before parsing
    expanded_items = []
    for level, text in items:
        if level == LEVEL_CONTENT and re.match(r'^art[\.\,]\s*\d+\s*[°ºª]?\s+a\s+\d+', text, re.IGNORECASE):
            expanded = expand_art_range_line(text)
            for exp_line in expanded:
                expanded_items.append((LEVEL_CONTENT, exp_line))
        else:
            expanded_items.append((level, text))
    items = expanded_items

    # Step 2: Parse entries
    print(f"\n[2/5] Parsing entries...")
    entries = parse_entries(items)

    main_entries = [e for e in entries if not e['is_cross_ref_only']]
    cross_refs = [e for e in entries if e['is_cross_ref_only']]

    total_arts = sum(
        len(art)
        for e in entries
        for sub in e['sub_entries']
        for art in [sub['articles']]
    )

    print(f"  Found {len(main_entries)} main entries, {len(cross_refs)} cross-reference entries")
    print(f"  Total article reference lines: {total_arts}")

    # Step 3: Convert to spreadsheet rows
    print(f"\n[3/5] Converting to spreadsheet format...")
    new_rows = entries_to_rows(entries)
    print(f"  Generated {len(new_rows)} spreadsheet rows")

    # Show some samples
    print("\n  Sample rows:")
    for i, (a, s, d, v) in enumerate(new_rows[:10]):
        disp_preview = d[:60] + '...' if len(d) > 60 else d
        print(f"    {a} | {s} | {disp_preview} | {v[:40] if v else ''}")

    # Step 4: Backup and merge
    print(f"\n[4/5] Creating backup at {BACKUP_PATH}...")
    shutil.copy2(str(XLSX_PATH), str(BACKUP_PATH))

    print(f"\n[5/5] Merging into spreadsheet...")
    appended, new_subs, new_assuntos = merge_into_xlsx(XLSX_PATH, new_rows, XLSX_PATH)

    print(f"\n{'=' * 60}")
    print(f"DONE!")
    print(f"  Existing entries complemented: {appended}")
    print(f"  New sub-assuntos added: {new_subs}")
    print(f"  New assuntos added: {new_assuntos}")
    print(f"  Backup saved to: {BACKUP_PATH}")
    print(f"  Updated file: {XLSX_PATH}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
