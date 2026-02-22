#!/usr/bin/env python3
"""
build_index.py - Parse the "indice remissivo" PDF and write entries to Excel.

Reads 'indice remissivo biblioteca.pdf', parses subject/sub-entry/dispositivos/vides,
converts reference notation, and writes to 'remissivo Teste.xlsx' preserving the
Normas sheet.

Usage:
    python build_index.py
"""

import re
import sys
import io
from pathlib import Path

import pdfplumber
import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PDF_PATH = SCRIPT_DIR / "índice remissivo biblioteca.pdf"
XLSX_PATH = SCRIPT_DIR / "remissivo Teste.xlsx"

# Header pattern present at the top of every page (page number varies)
HEADER_RE = re.compile(r"^REGIMENTO INTERNO DA CÂMARA MUNICIPAL DE SÃO PAULO \d+$")

# Patterns for "Ver:" and "Ver tambem:"
VER_RE = re.compile(r"^Ver:\s*(.+)$")
VER_TAMBEM_RE = re.compile(r"^Ver também:\s*(.+)$")

# Pattern for dispositivos in parentheses on a standalone line
DISPOSITIVOS_PAREN_RE = re.compile(r"^\(([^)]+)\)$")

# Pattern to detect an all-caps subject heading
SUBJECT_CHARS_RE = re.compile(r"[A-ZÀ-ÖØ-Ý]")
LOWER_CHARS_RE = re.compile(r"[a-zà-öø-ý]")

ROMAN_RE = re.compile(r"^[IVXLCDM]+$")


# ---------------------------------------------------------------------------
# User-validated existing entries (keep exactly as-is)
# ---------------------------------------------------------------------------

EXISTING_ENTRIES = [
    ('Proposições', '', '211-275', 'Emenda\nIndicação\nMoção\nProjeto de Lei\nRequerimento\nSubstitutivo'),
    ('Proposições', 'Adiamento de discussão ou votação', '175,II\n176\n192,II\n225,II\n285', ''),
    ('Proposições', 'Adiamento em bloco', '176,§10\nLOM:41,I', ''),
    ('Requerimento', '', '212', ''),
    ('Propositura', '', '', 'Projeto'),
    ('ABASTECIMENTO', 'Parecer da Comissão de Trânsito, Transporte e Atividade Econômica', '47,V,a,4', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Convocação dos responsáveis', '46,VI', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Criação de cargos e funções', '103,I,d\n105,XXVI', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Criação e organização', '47,IV,a,1', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Extinção de cargos', '105,XXVI', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Fiscalização dos atos', '46,IX', ''),
    ('ADMINISTRAÇÃO DIRETA E INDIRETA', 'Remuneração', '105,XXVI', ''),
    ('ADOLESCENTE', 'Proteção', '47,VII,a,4', ''),
    ('ANISTIA', 'Deliberação por maioria absoluta', '103,I,s', ''),
    ('APARTE', '', '139,IV\n306', ''),
    ('APARTE', 'Definição', '282', ''),
    ('APARTE', 'Duração', '282', ''),
    ('APARTE', 'Em audiência pública', '86,§2', ''),
    ('APARTE', 'Na Tribuna Popular', '209,§1', ''),
    ('APARTE', 'Não permitido', '23\n157\n283\n290,PU\n304\n309\n383,§1\n383,§2', ''),
    ('APARTE', 'Revisão', '150,PU\n283,§3', ''),
    ('Aplicação Financeira', '', '', 'CÂMARA MUNICIPAL|Aplicação Financeira'),
    ('Arma', '', '', 'PORTE DE ARMA'),
    ('ATA DAS SESSÕES', '', '146\n147', ''),
    ('ATA DAS SESSÕES', 'Aprovação', '147', ''),
    ('ATA DAS SESSÕES', 'Discussão', '147', ''),
    ('ATA DAS SESSÕES', 'Impugnação', '147,§1\n147,§3\n147,§7', ''),
    ('ATA DAS SESSÕES', 'Inscrição de voto de pesar', '223,VIII', ''),
    ('ATA DAS SESSÕES', 'Notas taquigráficas', '146', ''),
    ('ATA DAS SESSÕES', 'Publicação', '146\n147\n148\n149\n150\n183-A,§11', ''),
    ('ATA DAS SESSÕES', 'Redação', '26,VI\n26,VII', ''),
    ('ATA DAS SESSÕES', 'Retificação', '147,§1\n147,§3\n147,§7\n223,II', ''),
    ('ATA DAS SESSÕES', 'Revisão de discurso', '149\n150', ''),
    ('ATA DAS SESSÕES', 'Sessão secreta', '26,VII\n199\n200', ''),
]

# Build a set of (assunto, subassunto) keys for quick lookup
EXISTING_KEYS = set()
for _row in EXISTING_ENTRIES:
    EXISTING_KEYS.add((_row[0], _row[1]))


# ---------------------------------------------------------------------------
# Reference conversion functions
# ---------------------------------------------------------------------------

def convert_dispositivos(raw: str) -> str:
    """Convert a raw dispositivos string from the PDF into spreadsheet format.

    Examples:
        'arts.139,IV;306'            -> '139,IV\\n306'
        'art.86,§2º'                -> '86,§2'
        'arts.85 a 88;320;321'      -> '85-88\\n320\\n321'
        'art.147,§§1º,3º,7º'       -> '147,§1\\n147,§3\\n147,§7'
        'arts.383,§§1º e 2º'       -> '383,§1\\n383,§2'
        'art.26,VI,VII'             -> '26,VI\\n26,VII'
        'arts.150,p.único;283,§3º'  -> '150,PU\\n283,§3'
        'ADT art. 4ºC'             -> '4-C'
    """
    if not raw or not raw.strip():
        return ''

    raw = raw.strip()

    # Normalize whitespace (line wraps from PDF)
    raw = re.sub(r'\s+', ' ', raw).strip()

    # Handle quoted alinea letter patterns: "d" e "e" -> d,e (comma-separated alineas)
    # This way the alinea expansion logic can handle them: 17,V,d,e -> 17,V,d\n17,V,e
    raw = re.sub(r'"([a-z])"\s+e\s+"([a-z])"', r'\1,\2', raw)
    # Strip any remaining surrounding quotes from alinea letters: "e" -> e
    raw = re.sub(r'"([a-z])"', r'\1', raw)

    # Remove leading "art."/"arts." prefix
    raw = re.sub(r'^arts?\.\s*', '', raw)

    # Handle "e ADT art." at end of ref (before semicolon split)
    # e.g. "151 e ADT art. 4ºA" -> "151; ADT art. 4ºA"
    raw = re.sub(r'\s+e\s+(ADT\s+art)', r'; \1', raw)

    # Handle ",art." mid-reference as separator (not a sub-component)
    # e.g. "146,art.183-A,§11" -> "146;art.183-A,§11"
    # and "347,art.183-A" -> "347;art.183-A"
    raw = re.sub(r',\s*arts?\.\s*', ';', raw)

    # Handle comma after a single alinea letter followed by a new article number:
    # e.g. "17, V, e, 239,§1º" -> "17, V, e; 239,§1º"
    raw = re.sub(r',\s*([a-z])\s*,\s*(\d{2,})', r',\1;\2', raw)

    # Normalize "§ N" -> "§N" (remove space between § and digit)
    raw = re.sub(r'§\s+(\d)', r'§\1', raw)

    # Split on semicolons (each is a separate reference group)
    groups = [g.strip() for g in raw.split(';') if g.strip()]

    # Track ADT context: after an "ADT art." group, subsequent groups that
    # look like ADT article numbers (NºLETTER or e NºLETTER) inherit the ADT prefix.
    result_lines = []
    in_adt_context = False
    for group in groups:
        # Check if this group starts ADT context
        if re.match(r'ADT\s+arts?\.\s*', group, re.IGNORECASE):
            in_adt_context = True
            lines = convert_single_group(group)
            result_lines.extend(lines)
            continue

        # If in ADT context, check if this group is an ADT-like reference
        if in_adt_context:
            # Strip leading "e " if present: "e 4ºG" -> "4ºG"
            cleaned = re.sub(r'^e\s+', '', group).strip()
            if re.match(r'\d+º?[A-Z]', cleaned):
                # This is an ADT reference without explicit prefix
                lines = convert_adt_refs(cleaned)
                result_lines.extend(lines)
                continue
            else:
                # Not an ADT ref, exit context
                in_adt_context = False

        lines = convert_single_group(group)
        result_lines.extend(lines)

    return '\n'.join(result_lines)


def convert_single_group(group: str) -> list:
    """Convert a single reference group (no semicolons) into one or more output lines."""
    group = group.strip()
    if not group:
        return []

    # Remove any leading "art."/"arts." that might remain
    group = re.sub(r'^arts?\.\s*', '', group)

    # Check if this group contains an embedded ADT reference after a comma
    # e.g. "79, ADT art.4ºF" -> split into "79" and "ADT art.4ºF"
    adt_embedded = re.match(r'^(.+?),\s*(ADT\s+arts?\.\s*.+)$', group, re.IGNORECASE)
    if adt_embedded:
        left_part = adt_embedded.group(1).strip()
        adt_part = adt_embedded.group(2).strip()
        results = convert_single_group(left_part)
        results.extend(convert_single_group(adt_part))
        return results

    # Handle ADT references: "ADT art. 4ºC" -> "4-C"
    adt_match = re.match(r'ADT\s+arts?\.\s*(.+)', group, re.IGNORECASE)
    if adt_match:
        return convert_adt_refs(adt_match.group(1))

    # Handle "X a Y" range pattern at start: "85 a 88" -> "85-88"
    range_match = re.match(
        r'^(\d+(?:º)?(?:-[A-Z])?\s+a\s+\d+(?:º)?(?:-[A-Z])?(?:DT)?)(.*)$', group
    )
    if range_match:
        range_part = range_match.group(1)
        rest = range_match.group(2).strip()
        range_clean = re.sub(r'º', '', range_part).replace(' a ', '-')
        if rest:
            return [range_clean + rest]
        return [range_clean]

    # Handle "ao" as range: "9º ao 11" -> "9-11"
    ao_match = re.match(r'^(\d+)º?\s+ao\s+(\d+)(.*)$', group)
    if ao_match:
        result = f'{ao_match.group(1)}-{ao_match.group(2)}'
        rest = ao_match.group(3).strip()
        return [result + rest] if rest else [result]

    # Handle "X e Y" at top level (two separate articles joined by "e")
    # Only when both parts start with digits and it's not inside §§
    e_split = try_split_on_e_top_level(group)
    if e_split:
        result = []
        for part in e_split:
            result.extend(convert_single_group(part))
        return result

    # Now handle the main reference (article, components...)
    return expand_reference(group)


def convert_adt_refs(raw: str) -> list:
    """Convert ADT article references.
    '4ºC' -> '4-C', '4ºA e 4ºB' -> ['4-A', '4-B']
    '4ºD a 4ºF' -> '4-D-4-F'
    '4ºH, p.único' -> '4-H,PU'
    '4ºF, § 3º' -> '4-F,§3'
    """
    raw = raw.strip()

    # Handle range: "4ºD a 4ºF"
    range_match = re.match(r'(\d+)º?([A-Z])\s+a\s+(\d+)º?([A-Z])', raw)
    if range_match:
        return ['{}-{}-{}-{}'.format(
            range_match.group(1), range_match.group(2),
            range_match.group(3), range_match.group(4)
        )]

    # Split on " e " only (not on commas, which can be part of qualifiers like "4ºH, p.único")
    # But be careful: "4ºA e 4ºB" should split, while "4ºH, p.único" should not.
    parts = re.split(r'\s+e\s+', raw)
    results = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Match "4ºC" or "4ºC, §3º" or "4ºH, p.único"
        m = re.match(r'(\d+)º?([A-Z])(?:\s*,\s*(.+))?', part)
        if m:
            base = f'{m.group(1)}-{m.group(2)}'
            rest = m.group(3)
            if rest:
                results.append(f'{base},{clean_ref_component(rest)}')
            else:
                results.append(base)
        else:
            # Try range: "1º a 6ºDT"
            m2 = re.match(r'(\d+)º?\s+a\s+(\d+)º?(DT)?', part)
            if m2:
                results.append(f'{m2.group(1)}-{m2.group(2)}{m2.group(3) or ""}')
            else:
                results.append(clean_ref_component(part))
    return results


def try_split_on_e_top_level(group: str) -> list | None:
    """Split on ' e ' at the top level only when both parts are separate article refs.
    Returns None if splitting doesn't apply."""
    # Don't split inside §§ expansions
    if '§§' in group:
        return None

    # Don't split if left side contains § (the "e" connects § refs of same article)
    # e.g. "38,§1º e 3º" should NOT split (means 38,§1 and 38,§3)
    if '§' in group:
        return None

    m = re.match(r'^(.+?)\s+e\s+(\d+.*)$', group)
    if not m:
        return None

    left = m.group(1).strip()
    right = m.group(2).strip()

    # Both must start with digits
    if not re.match(r'^\d', left) or not re.match(r'^\d', right):
        return None

    # If left ends with a component ref (inciso/alinea after comma), don't split
    # e.g. "52,II,III" should NOT be split on "e"
    if re.search(r',\s*[IVXLCDMa-z]+\s*$', left):
        return None

    return [left, right]


def expand_reference(group: str) -> list:
    """Expand a single reference group. Handles §§, multiple incisos, etc."""
    group = group.strip()
    if not group:
        return []

    # Remove º after digits (but keep -A patterns like 183-A)
    # Handle ºDT -> DT, ºC -> C for ADT-like patterns already handled above
    group = re.sub(r'(\d)º(?=[A-Z])', r'\1', group)  # 4ºC -> 4C
    group = re.sub(r'º', '', group)

    # Replace p.único / par. único with PU
    group = re.sub(r'par\.\s*único', 'PU', group)
    group = re.sub(r'p\.\s*único', 'PU', group)
    group = re.sub(r'p\.único', 'PU', group)

    # Handle §§ (multiple paragraphs to expand)
    # Pattern: BASE,§§N1, N2, ... e Nk  where BASE can include inciso: "125,II" or just "383"
    ss_match = re.match(r'^(.+?)\s*,\s*§§\s*(.+)$', group)
    if ss_match:
        return expand_double_section(ss_match.group(1), ss_match.group(2))

    # Handle single § with "e" pattern: "38,§1 e 3" -> "38,§1\n38,§3"
    # This is like §§ but written with single § and "e"
    # BASE can include inciso components: "223,XV,§1 e 2"
    single_s_e = re.match(r'^(.+?)\s*,\s*§(\d+)\s+e\s+(\d+)$', group)
    if single_s_e:
        base = single_s_e.group(1)
        p1 = single_s_e.group(2)
        p2 = single_s_e.group(3)
        return [f'{base},§{p1}', f'{base},§{p2}']

    # Handle " a " range
    range_match = re.match(r'^(\d+(?:-[A-Z])?)\s+a\s+(\d+(?:-[A-Z])?)(.*)$', group)
    if range_match:
        start, end, rest = range_match.group(1), range_match.group(2), range_match.group(3).strip()
        return [f'{start}-{end}{rest}'] if rest else [f'{start}-{end}']

    # Handle multiple roman numerals with "e" at end:
    # "50,VII,VIII e IX" -> "50,VII\n50,VIII\n50,IX"
    multi_roman_e = try_expand_incisos_with_e(group)
    if multi_roman_e:
        return multi_roman_e

    # Handle multiple roman numerals at end: "26,VI,VII" -> "26,VI\n26,VII"
    multi = try_expand_multiple_incisos(group)
    if multi:
        return multi

    # Handle multiple roman numerals where last has qualifiers:
    # "223,VIII,XIV,§2" -> "223,VIII\n223,XIV,§2"
    multi_qual = try_expand_incisos_with_qualifier(group)
    if multi_qual:
        return multi_qual

    # Handle multiple alineas: "13,II,b,c" -> "13,II,b\n13,II,c"
    multi = try_expand_multiple_alineas(group)
    if multi:
        return multi

    # Clean spaces around commas
    group = re.sub(r'\s*,\s*', ',', group).strip()

    return [group]


def expand_double_section(base: str, paras_raw: str) -> list:
    """Expand §§ references.
    base='383', paras_raw='1, 2, 3 e 4' -> ['383,§1', '383,§2', '383,§3', '383,§4']
    base='125,II', paras_raw='1,2' -> ['125,II,§1', '125,II,§2']
    """
    # Clean base: remove extra spaces around commas
    base = re.sub(r'\s*,\s*', ',', base).strip()
    paras_raw = re.sub(r'º', '', paras_raw)
    # Replace " e " with ","
    paras_raw = re.sub(r'\s+e\s+', ',', paras_raw)
    parts = [p.strip() for p in paras_raw.split(',') if p.strip()]
    return [f'{base},§{p}' for p in parts]


def is_roman(s: str) -> bool:
    """Check if a string looks like a Roman numeral."""
    return bool(ROMAN_RE.match(s.strip()))


def try_expand_incisos_with_e(group: str) -> list | None:
    """Expand roman numerals joined with 'e' at the end.
    '50,VII,VIII e IX' -> ['50,VII', '50,VIII', '50,IX']
    '225,VI,VII' already handled by try_expand_multiple_incisos.
    This handles the case where the last item uses 'e' instead of comma."""
    # Match: base,ROMAN1,ROMAN2 e ROMAN3  or  base,ROMAN1 e ROMAN2
    m = re.match(r'^(.+?),\s*([IVXLCDM]+(?:\s*,\s*[IVXLCDM]+)*)\s+e\s+([IVXLCDM]+)$', group)
    if not m:
        return None
    base = m.group(1).strip()
    roman_list_str = m.group(2).strip()
    last_roman = m.group(3).strip()

    # Split the roman numerals
    romans = [r.strip() for r in roman_list_str.split(',')]
    romans.append(last_roman)

    # Verify all are roman
    if not all(is_roman(r) for r in romans):
        return None

    return [f'{base},{r}' for r in romans]


def try_expand_multiple_incisos(group: str) -> list | None:
    """Expand when last two comma-separated components are both Roman numerals.
    '26,VI,VII' -> ['26,VI', '26,VII']"""
    parts = group.split(',')
    if len(parts) < 3:
        return None
    last = parts[-1].strip()
    second_last = parts[-2].strip()
    if is_roman(last) and is_roman(second_last):
        base = ','.join(p.strip() for p in parts[:-2])
        return [f'{base},{second_last}', f'{base},{last}']
    return None


def try_expand_incisos_with_qualifier(group: str) -> list | None:
    """Expand when there are two consecutive roman numerals but the last has qualifiers.
    '223,VIII,XIV,§2' -> ['223,VIII', '223,XIV,§2']
    The pattern is: ART,ROMAN1,ROMAN2,QUALIFIER where ROMAN1 and ROMAN2 are roman,
    and QUALIFIER starts with § or is PU, etc."""
    parts = [p.strip() for p in group.split(',')]
    if len(parts) < 4:
        return None
    # We need: parts[0]=article, at least two roman numerals in middle, then qualifier on last roman
    # Find the article number (first part that starts with a digit)
    if not re.match(r'^\d', parts[0]):
        return None
    # Find consecutive roman numeral parts
    # Walk from index 1 looking for the first roman numeral
    first_roman_idx = None
    for k in range(1, len(parts)):
        if is_roman(parts[k]):
            first_roman_idx = k
            break
    if first_roman_idx is None:
        return None
    # Check if there's a second roman numeral
    second_roman_idx = None
    for k in range(first_roman_idx + 1, len(parts)):
        if is_roman(parts[k]):
            second_roman_idx = k
            break
    if second_roman_idx is None:
        return None
    # The parts after the second roman (if any) are qualifiers for the second roman
    # The first roman stands alone
    base = ','.join(parts[:first_roman_idx])
    first_inciso = parts[first_roman_idx]
    second_inciso_with_qualifier = ','.join(parts[second_roman_idx:])
    return [f'{base},{first_inciso}', f'{base},{second_inciso_with_qualifier}']


def try_expand_multiple_alineas(group: str) -> list | None:
    """Expand when last two components are single lowercase letters.
    '13,II,b,c' -> ['13,II,b', '13,II,c']"""
    parts = group.split(',')
    if len(parts) < 4:
        return None
    last = parts[-1].strip()
    second_last = parts[-2].strip()
    if re.match(r'^[a-z]$', last) and re.match(r'^[a-z]$', second_last):
        base = ','.join(p.strip() for p in parts[:-2])
        return [f'{base},{second_last}', f'{base},{last}']
    return None


def clean_ref_component(s: str) -> str:
    """Clean a reference component: remove º, convert p.único -> PU, etc."""
    s = re.sub(r'º', '', s)
    s = re.sub(r'par\.\s*único', 'PU', s)
    s = re.sub(r'p\.\s*único', 'PU', s)
    s = re.sub(r'p\.único', 'PU', s)
    s = re.sub(r'\s*,\s*', ',', s)
    return s.strip()


# ---------------------------------------------------------------------------
# PDF Parsing
# ---------------------------------------------------------------------------

def extract_all_lines(pdf_path: Path) -> list:
    """Extract all lines from the PDF, skipping page headers and INDICE title."""
    all_lines = []
    pdf = pdfplumber.open(str(pdf_path))

    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
            if HEADER_RE.match(line):
                continue
            if line == 'ÍNDICE':
                continue
            all_lines.append(line)

    pdf.close()
    return all_lines


def is_all_caps_line(line: str) -> bool:
    """Check if a line is an all-caps subject heading.
    Subject headings have all alphabetic chars in uppercase and at least 2 letters."""
    upper_count = len(SUBJECT_CHARS_RE.findall(line))
    lower_count = len(LOWER_CHARS_RE.findall(line))
    return upper_count >= 2 and lower_count == 0


def join_wrapped_lines(lines: list) -> list:
    """Pre-process: join continuation lines that wrap from the previous line.

    A continuation line is one that:
    - Does NOT start with '- ' (sub-entry)
    - Does NOT match 'Ver:' or 'Ver também:'
    - Does NOT match a standalone parenthetical dispositivos '(arts...)'
    - Is NOT an all-caps line (potential subject heading)
    - Appears to be a continuation of a sub-entry or parenthetical that wrapped

    We detect this by checking if the line starts with a lowercase letter, digit,
    or special char that continues the previous line's content.
    """
    result = []
    i = 0
    while i < len(lines):
        line = lines[i]

        # Check if the NEXT line should be joined to this one
        while i + 1 < len(lines):
            next_line = lines[i + 1]

            # Is next_line a continuation?
            if is_continuation_line(next_line, line):
                # Join: add a space between them
                line = line + ' ' + next_line
                i += 1
            else:
                break

        result.append(line)
        i += 1

    return result


def is_continuation_line(line: str, prev_line: str) -> bool:
    """Determine if 'line' is a continuation of 'prev_line' (wrapped text).

    Continuation lines are those that don't start their own structural element
    and instead continue text from the previous line. Common cases:
    - A sub-entry's dispositivos wrapped to the next line
    - A parenthetical reference that wrapped
    - A sub-entry text that was too long
    """
    # Never join if line starts a new structural element
    if line.startswith('- '):
        return False
    if VER_RE.match(line) or VER_TAMBEM_RE.match(line):
        return False
    if DISPOSITIVOS_PAREN_RE.match(line):
        return False

    # All-caps lines are subject headings (not continuations), UNLESS
    # prev_line ends with a preposition indicating it's a multi-line heading
    # (but we handle multi-line headings separately in the parser)
    if is_all_caps_line(line):
        return False

    # If previous line was a sub-entry or parenthetical that ended with a comma,
    # semicolon, or open paren, this is a continuation
    if prev_line.endswith(',') or prev_line.endswith('('):
        return True

    # If previous line starts with '- ' and this line continues the content
    # (e.g. wrapped dispositivos ending with ')')
    if prev_line.startswith('- ') and line.rstrip().endswith(')'):
        return True

    # If previous line starts with '- ' and doesn't end with ')' but the reference
    # was split (the line contains reference patterns like 'IV;381 a 384)')
    if prev_line.startswith('- ') and re.match(r'^[A-Z0-9§,;.\s()ºúa-z]+$', line):
        # Looks like continuation of a reference or text
        # Check if it has reference-like content
        if re.search(r'[0-9§;,]', line):
            return True

    # If previous line is a parenthetical that wraps (started with '(' but doesn't close)
    if prev_line.startswith('(') and prev_line.count('(') > prev_line.count(')'):
        return True

    # If prev_line is a "Ver:" line and this line starts with "- " ... no, already handled.
    # If prev_line is a sub-entry that ends without closing paren but had an opening paren
    if '(' in prev_line and prev_line.count('(') > prev_line.count(')'):
        return True

    return False


def next_content_line_type(lines: list, start_idx: int) -> str:
    """Look ahead from start_idx to determine what type of content follows.
    Returns: 'sub_entry', 'parenthetical', 'ver', 'ver_tambem', 'subject', 'end'"""
    for j in range(start_idx, len(lines)):
        line = lines[j]
        if line.startswith('- '):
            return 'sub_entry'
        if DISPOSITIVOS_PAREN_RE.match(line):
            return 'parenthetical'
        if VER_RE.match(line):
            return 'ver'
        if VER_TAMBEM_RE.match(line):
            return 'ver_tambem'
        if is_all_caps_line(line):
            return 'subject'
    return 'end'


def is_new_subject_after_vide(lines: list, idx: int) -> bool:
    """Check if the all-caps line at 'idx' is a new subject rather than a vide continuation.

    An all-caps line is a new subject if the FIRST NON-CONTINUATION line after it is:
    - A sub-entry '- ...'
    - A parenthetical dispositivos '(arts...)'
    - A 'Ver:' or 'Ver também:' line
    - Or if there is no next line (end of index)

    An all-caps line is a vide continuation if the first non-continuation line after it is:
    - Another all-caps line (another vide target that is NOT a multi-line continuation)

    Multi-line subject headings (like commission names) require looking past continuation
    lines to see what structural element follows the complete heading.
    """
    # Skip past the candidate line and any multi-line continuation lines
    j = idx + 1
    accumulated = lines[idx]
    while j < len(lines):
        line = lines[j]
        if is_all_caps_line(line) and looks_like_subject_continuation(accumulated, line, strict=True):
            accumulated = accumulated + ' ' + line
            j += 1
        else:
            break

    if j >= len(lines):
        return True

    # Now check what follows the (possibly multi-line) heading
    first_after = lines[j]

    if first_after.startswith('- '):
        return True
    if DISPOSITIVOS_PAREN_RE.match(first_after):
        return True
    if VER_RE.match(first_after) or VER_TAMBEM_RE.match(first_after):
        return True

    # Otherwise treat as vide continuation
    return False


def looks_like_subject_continuation(current: str, next_line: str, strict: bool = False) -> bool:
    """Determine if next_line continues a multi-line subject heading.
    Returns True if 'current' ends with a word that indicates continuation.

    When strict=True (used for vide targets), only use the preposition/article
    heuristic, not the COMISSÃO special case. This prevents merging unrelated
    vide targets that happen to be commission names.
    """
    continuation_endings = {
        'DA', 'DO', 'DOS', 'DAS', 'DE', 'E', 'EM', 'AO', 'À',
        'NO', 'NA', 'NOS', 'NAS', 'POR', 'PARA', 'COM', 'SEM',
    }

    last_word = current.rstrip().split()[-1].upper() if current.strip() else ''

    if last_word in continuation_endings:
        return True

    # Words that typically continue a commission name (adjective/modifier)
    # rather than starting a new subject heading.
    continuation_starters = {
        'HUMANOS', 'SOCIAL', 'INTERNACIONAIS', 'INTELIGENTE',
        'EVENTOS', 'CIDADE', 'ANIMAIS',
    }
    first_word_next = next_line.split()[0].upper() if next_line.strip() else ''

    if strict:
        # In strict mode, also allow continuation_starters when the accumulated
        # text is a commission name (starts with COMISSÃO). This handles vide
        # targets that are long commission names spanning multiple lines.
        if 'COMISSÃO' in current.upper() and first_word_next in continuation_starters:
            return True
        return False

    # For multi-line subject headings (non-strict mode), also check if the
    # current line ends with a word that is part of a multi-word name that
    # breaks across lines. We use a targeted pattern: if current line
    # ends with a noun like DIREITOS, ASSISTÊNCIA, RELAÇÕES, and the next
    # line starts with a modifier (adjective) that completes the name.
    # These patterns occur in commission names.
    if first_word_next in continuation_starters:
        return True

    return False


def parse_sub_entry(text: str) -> tuple:
    """Parse a sub-entry text into (subassunto_text, raw_dispositivos).
    'Definição (art.282)' -> ('Definição', 'art.282')
    'Do Prefeito (arts.47,II,a;105,XII;295, p.único,II;385' -> ('Do Prefeito', 'arts.47,II,a;...')
    """
    if text.endswith(')'):
        depth = 0
        for j in range(len(text) - 1, -1, -1):
            if text[j] == ')':
                depth += 1
            elif text[j] == '(':
                depth -= 1
                if depth == 0:
                    sub_text = text[:j].strip()
                    disp_raw = text[j + 1:-1].strip()
                    # Verify it looks like a dispositivos reference
                    if re.search(r'(?:arts?\.|^\d|§|ADT)', disp_raw):
                        return (sub_text, disp_raw)
                    else:
                        return (text, '')
                    break

    # Handle unclosed parenthesis: "Do Prefeito (arts.47,II,a;105,XII;295, p.único,II;385"
    # If text has a '(' but no closing ')' and the part after '(' looks like a reference
    if '(' in text and text.count('(') > text.count(')'):
        paren_idx = text.index('(')
        sub_text = text[:paren_idx].strip()
        disp_raw = text[paren_idx + 1:].strip()
        # Remove trailing whitespace and any trailing unclosed part
        disp_raw = disp_raw.rstrip()
        if re.search(r'(?:arts?\.|^\d|§|ADT)', disp_raw):
            return (sub_text, disp_raw)

    # Some sub-entries have no dispositivos (just text)
    return (text, '')


def convert_ver_target(target: str) -> str:
    """Convert a Ver/Ver também target.
    'CÂMARA MUNICIPAL - Aplicação Financeira' -> 'CÂMARA MUNICIPAL|Aplicação Financeira'
    """
    return re.sub(r'\s+-\s+', '|', target, count=1)


def parse_entries(lines: list) -> list:
    """Parse lines into a list of (assunto, subassunto, raw_dispositivos, vides) entries.

    The parser uses a state machine with look-ahead to correctly handle:
    - Multi-line subject names
    - Ver/Ver também with multiple continuation targets
    - Page breaks mid-entry
    - Parenthetical dispositivos on the subject's own line
    """
    entries = []
    current_subject = None
    current_subject_dispositivos = None
    current_subject_vides = []
    current_sub_entries = []  # list of (subassunto_text, raw_dispositivos)

    # State for collecting Ver/Ver também targets
    in_ver_tambem = False
    ver_tambem_targets = []
    pending_ver_target = None  # A "Ver:" that might get a sub-qualifier

    # State for collecting multiple Ver: targets (like CONVOCAÇÃO EXTRAORDINÁRIA)
    in_ver_continuation = False

    def flush_subject():
        """Save the current subject and all its entries."""
        nonlocal current_subject, current_subject_dispositivos, current_subject_vides
        nonlocal current_sub_entries, in_ver_tambem, ver_tambem_targets
        nonlocal pending_ver_target, in_ver_continuation

        if current_subject is None:
            return

        flush_ver_state()

        # Build vides string
        vides_str = '\n'.join(current_subject_vides) if current_subject_vides else ''

        # Create main entry if it has dispositivos, vides, or no sub-entries
        if current_subject_dispositivos or vides_str or not current_sub_entries:
            entries.append((
                current_subject,
                '',
                current_subject_dispositivos or '',
                vides_str
            ))

        # Add sub-entries
        for sub_text, sub_disp in current_sub_entries:
            entries.append((current_subject, sub_text, sub_disp, ''))

        # Reset state
        current_subject = None
        current_subject_dispositivos = None
        current_subject_vides = []
        current_sub_entries = []
        in_ver_tambem = False
        ver_tambem_targets = []
        pending_ver_target = None
        in_ver_continuation = False

    def flush_ver_state():
        """Flush any pending Ver/Ver também state into vides."""
        nonlocal in_ver_tambem, ver_tambem_targets, pending_ver_target, in_ver_continuation
        if ver_tambem_targets:
            current_subject_vides.extend(ver_tambem_targets)
            ver_tambem_targets = []
        if pending_ver_target:
            current_subject_vides.append(pending_ver_target)
            pending_ver_target = None
        in_ver_tambem = False
        in_ver_continuation = False

    i = 0
    while i < len(lines):
        line = lines[i]

        # --- Sub-entry lines ---
        if line.startswith('- '):
            # Check if this is a sub-qualifier for a pending Ver: target
            # e.g. "Ver: SITUAÇÃO DE EMERGÊNCIA" + "- De saúde pública"
            if pending_ver_target is not None:
                sub_text = line[2:].strip()
                # Don't include dispositivos from the sub-qualifier
                sub_text_clean, _ = parse_sub_entry(sub_text)
                pending_ver_target = pending_ver_target + '|' + sub_text_clean
                # Flush immediately
                current_subject_vides.append(pending_ver_target)
                pending_ver_target = None
                in_ver_continuation = False
                i += 1
                continue

            # End any Ver/Ver também collection
            flush_ver_state()

            if current_subject is not None:
                sub_line = line[2:].strip()
                sub_text, sub_disp = parse_sub_entry(sub_line)
                current_sub_entries.append((sub_text, sub_disp))
            i += 1
            continue

        # --- Ver: lines ---
        ver_match = VER_RE.match(line)
        if ver_match:
            # Flush any previous Ver state
            if ver_tambem_targets:
                current_subject_vides.extend(ver_tambem_targets)
                ver_tambem_targets = []
            if pending_ver_target:
                current_subject_vides.append(pending_ver_target)

            in_ver_tambem = False

            target = convert_ver_target(ver_match.group(1).strip())
            pending_ver_target = target
            in_ver_continuation = True  # After Ver:, all-caps lines might be more targets
            i += 1
            continue

        # --- Ver também: lines ---
        ver_tambem_match = VER_TAMBEM_RE.match(line)
        if ver_tambem_match:
            # Flush pending Ver target
            if pending_ver_target:
                current_subject_vides.append(pending_ver_target)
                pending_ver_target = None
            in_ver_continuation = False

            in_ver_tambem = True
            target = convert_ver_target(ver_tambem_match.group(1).strip())
            ver_tambem_targets.append(target)
            i += 1
            continue

        # --- Standalone parenthetical dispositivos ---
        paren_match = DISPOSITIVOS_PAREN_RE.match(line)
        if paren_match:
            flush_ver_state()
            if current_subject is not None and current_subject_dispositivos is None:
                current_subject_dispositivos = paren_match.group(1)
            i += 1
            continue

        # --- All-caps lines (potential subjects or vide continuations) ---
        if is_all_caps_line(line):

            # CASE 1: We're in Ver também mode - is this a continuation vide or new subject?
            if in_ver_tambem:
                if is_new_subject_after_vide(lines, i):
                    # This is a new subject, not a vide
                    flush_ver_state()
                    flush_subject()
                    current_subject = consume_multiline_subject(lines, i)
                    i = skip_consumed_subject_lines(lines, i, current_subject)
                    continue
                else:
                    # It's a vide continuation target (use strict mode for multi-line)
                    target = consume_multiline_subject(lines, i, strict=True)
                    ver_tambem_targets.append(target)
                    i = skip_consumed_subject_lines(lines, i, target, strict=True)
                    continue

            # CASE 2: We have a pending Ver: target and this could be another Ver target
            if in_ver_continuation:
                # Flush the pending Ver target first
                if pending_ver_target:
                    current_subject_vides.append(pending_ver_target)
                    pending_ver_target = None

                if is_new_subject_after_vide(lines, i):
                    # New subject
                    in_ver_continuation = False
                    flush_subject()
                    current_subject = consume_multiline_subject(lines, i)
                    i = skip_consumed_subject_lines(lines, i, current_subject)
                    continue
                else:
                    # Additional vide target (use strict mode)
                    target = consume_multiline_subject(lines, i, strict=True)
                    current_subject_vides.append(target)
                    i = skip_consumed_subject_lines(lines, i, target, strict=True)
                    continue

            # CASE 3: Regular new subject heading
            flush_ver_state()
            flush_subject()

            current_subject = consume_multiline_subject(lines, i)
            i = skip_consumed_subject_lines(lines, i, current_subject)
            continue

        # --- Unrecognized line (shouldn't happen after line joining) ---
        # Flush any state and skip
        flush_ver_state()
        i += 1
        continue

    # Flush the last subject
    flush_subject()

    return entries


def consume_multiline_subject(lines: list, start: int, strict: bool = False) -> str:
    """Starting at 'start', consume all lines that form a multi-line subject heading.
    Returns the complete subject name.

    When strict=True (used for vide targets), only join on preposition/article endings,
    not on the COMISSÃO heuristic.
    """
    subject = lines[start]
    j = start + 1
    while j < len(lines):
        next_line = lines[j]
        if is_all_caps_line(next_line) and looks_like_subject_continuation(subject, next_line, strict=strict):
            subject = subject + ' ' + next_line
            j += 1
        else:
            break
    return subject


def skip_consumed_subject_lines(lines: list, start: int, subject: str, strict: bool = False) -> int:
    """Given that we consumed a (possibly multi-line) subject starting at 'start',
    return the index of the next unprocessed line."""
    j = start
    consumed = lines[start]
    j += 1
    while j < len(lines):
        next_line = lines[j]
        if is_all_caps_line(next_line) and looks_like_subject_continuation(consumed, next_line, strict=strict):
            consumed = consumed + ' ' + next_line
            j += 1
        else:
            break
    return j


# ---------------------------------------------------------------------------
# Post-processing: convert raw dispositivos to spreadsheet format
# ---------------------------------------------------------------------------

def process_entries(raw_entries: list) -> list:
    """Convert raw dispositivos in all entries to spreadsheet format."""
    processed = []
    for assunto, sub, raw_disp, vides in raw_entries:
        converted = convert_dispositivos(raw_disp) if raw_disp else ''
        processed.append((assunto, sub, converted, vides))
    return processed


# ---------------------------------------------------------------------------
# Merge with existing validated entries
# ---------------------------------------------------------------------------

def merge_entries(parsed_entries: list, existing: list) -> list:
    """Merge parsed entries with existing validated entries.
    Existing entries take priority. TESTE entries are removed."""
    # Build case-insensitive lookup of existing keys
    existing_upper_keys = set()
    for row in existing:
        existing_upper_keys.add((row[0].upper(), row[1].upper()))

    result = list(existing)

    for entry in parsed_entries:
        if entry[0] == 'TESTE':
            continue

        upper_key = (entry[0].upper(), entry[1].upper())
        if upper_key in existing_upper_keys:
            continue

        result.append(entry)

    return result


# ---------------------------------------------------------------------------
# Write to Excel
# ---------------------------------------------------------------------------

def write_xlsx(entries: list, xlsx_path: Path):
    """Write entries to Excel, preserving the Normas sheet."""
    wb = openpyxl.load_workbook(str(xlsx_path))

    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
        ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet('Sheet1', 0)
        ws.append(['Assunto', 'SubAssunto', 'Dispositivos', 'Vides'])

    for assunto, sub, disp, vides in entries:
        ws.append([
            assunto or None,
            sub or None,
            disp or None,
            vides or None,
        ])

    wb.save(str(xlsx_path))
    print(f"Wrote {len(entries)} entries to {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Ensure UTF-8 output on Windows console
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

    print(f"Reading PDF: {PDF_PATH}")
    if not PDF_PATH.exists():
        print(f"ERROR: PDF file not found: {PDF_PATH}")
        sys.exit(1)

    # Step 1: Extract all lines from PDF
    raw_lines = extract_all_lines(PDF_PATH)
    print(f"Extracted {len(raw_lines)} raw lines from PDF")

    # Step 2: Join wrapped/continuation lines
    lines = join_wrapped_lines(raw_lines)
    print(f"After joining continuations: {len(lines)} lines")

    # Step 3: Parse entries
    raw_entries = parse_entries(lines)
    print(f"Parsed {len(raw_entries)} raw entries")

    # Step 4: Convert dispositivos to spreadsheet format
    processed_entries = process_entries(raw_entries)
    print(f"Processed {len(processed_entries)} entries")

    # Step 5: Merge with existing validated entries
    merged = merge_entries(processed_entries, EXISTING_ENTRIES)
    print(f"Merged to {len(merged)} total entries (including {len(EXISTING_ENTRIES)} validated)")

    # Step 6: Write to Excel
    if not XLSX_PATH.exists():
        print(f"ERROR: Excel file not found: {XLSX_PATH}")
        sys.exit(1)

    write_xlsx(merged, XLSX_PATH)

    # Print summary
    subjects = set(e[0] for e in merged)
    print(f"\nSummary:")
    print(f"  Total entries: {len(merged)}")
    print(f"  Unique subjects: {len(subjects)}")
    print(f"  Existing (validated): {len(EXISTING_ENTRIES)}")
    print(f"  New from PDF: {len(merged) - len(EXISTING_ENTRIES)}")

    # Print first 15 entries for verification
    print(f"\nFirst 15 entries:")
    for i, (a, s, d, v) in enumerate(merged[:15]):
        d_preview = (d or '').replace('\n', '|')[:60]
        v_preview = (v or '').replace('\n', '|')[:60]
        print(f"  {i+1}. [{a}] [{s}] [{d_preview}] [{v_preview}]")


if __name__ == '__main__':
    main()
