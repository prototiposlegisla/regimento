"""Validação de remissivo.xlsx conforme instruções de preenchimento."""

from __future__ import annotations

import re
from pathlib import Path

_ROMAN_RE = re.compile(r"^[IVXLC]+$")
_ALINEA_RE = re.compile(r"^[a-z]$")
_ITEM_RE = re.compile(r"^\d+$")
_ART_RE = re.compile(r"^\d+[-A-Za-z]*$")
_RANGE_RE = re.compile(r"^\d+\s*[-\u2013\u2014]\s*\d+$")
_LAW_PREFIX_LINE_RE = re.compile(r"^([A-Za-z]{2,})\s*:\s*(.+)$")
_HINT_RE = re.compile(r"\(([^)]+)\)\s*$")


def _validate_detail(detail: str) -> str | None:
    """Retorna mensagem de erro se detalhe inválido, None se válido."""
    d = detail.strip()

    # caput / PU / §ú
    if d.lower() == "caput":
        return None
    if d.upper() == "PU" or d in ("\u00a7\u00fa", "\u00a7u"):
        return None
    if re.match(r"^\u00a7\d+$", d):
        return None

    parts = [p.strip() for p in d.split(",")]

    if len(parts) == 1:
        p = parts[0]
        if _ROMAN_RE.match(p) or _ALINEA_RE.match(p) or _ITEM_RE.match(p):
            return None
        return f"detalhe desconhecido: '{d}'"

    if len(parts) == 2:
        p0, p1 = parts
        if _ROMAN_RE.match(p0) and _ROMAN_RE.match(p1):
            return f"múltiplos incisos na mesma linha — use linhas separadas: '{p0}' e '{p1}'"
        if _ROMAN_RE.match(p0) and _ALINEA_RE.match(p1):
            return None  # inciso,alínea ✓
        if _ALINEA_RE.match(p0) and _ITEM_RE.match(p1):
            return None  # alínea,item ✓
        if _ALINEA_RE.match(p0) and _ALINEA_RE.match(p1):
            return f"múltiplas alíneas na mesma linha — use linhas separadas"
        # §ú,inciso ou §N,inciso (parágrafo com inciso) ✓
        if re.match(r"^§(\d+|ú|u)$", p0) and _ROMAN_RE.match(p1):
            return None
        return f"estrutura de detalhe inválida: '{d}'"

    if len(parts) == 3:
        p0, p1, p2 = parts
        if _ROMAN_RE.match(p0) and _ALINEA_RE.match(p1) and _ITEM_RE.match(p2):
            return None  # inciso,alínea,item ✓
        return f"estrutura de detalhe inválida (esperado: inciso,alínea,item): '{d}'"

    return f"detalhe com muitas partes: '{d}'"


def _validate_device_line(raw_line: str, known_prefixes: set[str]) -> list[str]:
    """Valida uma linha de dispositivo. Retorna lista de erros."""
    errors: list[str] = []
    line = raw_line.strip()
    if not line:
        return errors

    # 1. Espaço após vírgula
    if ", " in line:
        errors.append(
            f"espaço após vírgula — use '{line.replace(', ', ',')}' em vez de '{line}'"
        )
        line = line.replace(", ", ",")

    # 2. Espaço ao redor do ':' do prefixo de lei
    if re.match(r"^[A-Za-z]{2,}\s+:", line) or re.match(r"^[A-Za-z]{2,}:\s+", line):
        errors.append(
            f"espaço ao redor do ':' no prefixo — use 'SIGLA:artigo' sem espaços: '{raw_line.strip()}'"
        )

    # 3. Extrair prefixo de lei
    law_prefix = ""
    law_m = _LAW_PREFIX_LINE_RE.match(line)
    if law_m:
        law_prefix = law_m.group(1).upper()
        line = law_m.group(2).strip()
        # Validar que prefixo existe na aba Normas
        if known_prefixes and law_prefix not in known_prefixes:
            errors.append(
                f"prefixo '{law_prefix}' não cadastrado na aba Normas "
                f"(prefixos conhecidos: {', '.join(sorted(known_prefixes))})"
            )

    # 4. Extrair dica entre parênteses no fim
    hint_m = _HINT_RE.search(line)
    if hint_m:
        line = line[: hint_m.start()].strip()

    # 5. Range de artigos (ex: "211-275")
    if _RANGE_RE.match(line):
        # Garante que é realmente dois números e não algo como "4-A"
        nums = re.split(r"[-\u2013\u2014]", line)
        if len(nums) == 2 and nums[0].strip().isdigit() and nums[1].strip().isdigit():
            start, end = int(nums[0].strip()), int(nums[1].strip())
            if start >= end:
                errors.append(
                    f"range inválido: início ({start}) deve ser menor que fim ({end})"
                )
            return errors

    # 6. Artigo simples ou com detalhe
    if not line:
        errors.append("referência vazia após extração de prefixo/dica")
        return errors

    parts = line.split(",", 1)
    art = parts[0].strip()

    if not _ART_RE.match(art):
        errors.append(f"número de artigo inválido: '{art}'")
        return errors

    if len(parts) == 2:
        detail = parts[1].strip()
        if detail:
            err = _validate_detail(detail)
            if err:
                errors.append(err)

    return errors


def validate_xlsx(path: str | Path, law_mapping: dict[str, str]) -> list[str]:
    """Valida formato de remissivo.xlsx conforme instruções de preenchimento.

    Retorna lista de strings de erro/aviso para exibição no log do build.
    ``law_mapping`` deve ser o dict {nome_lei: prefixo} retornado por parse_law_mapping.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    messages: list[str] = []

    has_normas = "Normas" in wb.sheetnames
    if not has_normas:
        messages.append(
            "  aviso: aba 'Normas' ausente — prefixos de leis externas não serão validados"
        )

    # Prefixos conhecidos (valores do mapeamento)
    known_prefixes: set[str] = set(law_mapping.values())

    # Localiza aba principal
    ws = None
    for name in wb.sheetnames:
        if name != "Normas":
            ws = wb[name]
            break

    if ws is None:
        wb.close()
        messages.append("  erro: nenhuma aba principal (não-Normas) encontrada na planilha")
        return messages

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    for i, row in enumerate(rows, start=2):  # linha 1 = cabeçalho
        if not row or len(row) < 1:
            continue

        assunto = str(row[0] or "").strip()
        if not assunto:
            continue  # ignorado conforme especificação

        sub_assunto = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
        ctx = assunto + (f" > {sub_assunto}" if sub_assunto else "")

        # ── Coluna Dispositivos ─────────────────────────────────────
        if len(row) > 2 and row[2]:
            disp_raw = str(row[2]).strip()
            for ln, dline in enumerate(disp_raw.split("\n"), start=1):
                dline = dline.strip()
                if not dline:
                    continue
                for err in _validate_device_line(dline, known_prefixes):
                    messages.append(
                        f"  Linha {i} ({ctx}) · Dispositivos[{ln}] '{dline}': {err}"
                    )

        # ── Coluna Vide ─────────────────────────────────────────────
        if len(row) > 3 and row[3]:
            vide_raw = str(row[3]).strip()
            for ln, vline in enumerate(vide_raw.split("\n"), start=1):
                vline = vline.strip()
                if not vline:
                    continue
                pipe_count = vline.count("|")
                if pipe_count > 1:
                    messages.append(
                        f"  Linha {i} ({ctx}) · Vide[{ln}] '{vline}': "
                        f"múltiplos '|' — use exatamente um separador: 'Assunto|Sub-assunto'"
                    )
                elif pipe_count == 1:
                    parts = vline.split("|", 1)
                    if not parts[0].strip():
                        messages.append(
                            f"  Linha {i} ({ctx}) · Vide[{ln}] '{vline}': assunto vazio antes de '|'"
                        )
                    if not parts[1].strip():
                        messages.append(
                            f"  Linha {i} ({ctx}) · Vide[{ln}] '{vline}': sub-assunto vazio após '|'"
                        )

    return messages
