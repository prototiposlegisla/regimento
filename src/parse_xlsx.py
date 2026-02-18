"""Parser de remissivo.xlsx → SubjectIndex + mapeamento de normas."""

from __future__ import annotations

import re
from pathlib import Path

from .models import SubjectEntry, SubjectIndex, SubjectRef


def parse_law_mapping(path: str | Path) -> dict[str, str]:
    """Lê aba 'Normas' do XLSX → {nome: prefixo}.

    A aba deve ter colunas: Prefixo | Nome.
    Retorna dict vazio se a aba não existir.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    mapping: dict[str, str] = {}
    if "Normas" in wb.sheetnames:
        ws = wb["Normas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            prefix = str(row[0] or "").strip()
            name = str(row[1] or "").strip()
            if prefix and name:
                mapping[name] = prefix
    wb.close()
    return mapping


def parse_xlsx(path: str | Path, known_lettered: set[str] | None = None) -> SubjectIndex:
    """Parseia remissivo.xlsx e retorna SubjectIndex.

    known_lettered: conjunto de art_numbers letrados conhecidos (ex: {"212-A", "183-A"})
    para incluir em expansões de range.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Use first sheet that isn't "Normas" (avoid depending on active sheet)
    ws = None
    for name in wb.sheetnames:
        if name != "Normas":
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return SubjectIndex(entries=[])

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    entries: list[SubjectEntry] = []
    for row in rows:
        if not row or len(row) < 3:
            continue

        assunto = str(row[0] or "").strip()
        sub_assunto = str(row[1] or "").strip() if row[1] else ""
        dispositivos_raw = str(row[2] or "").strip() if row[2] else ""
        vides_raw = str(row[3] or "").strip() if len(row) > 3 and row[3] else ""

        if not assunto:
            continue

        refs = _parse_dispositivos(dispositivos_raw, known_lettered)
        vides = [v.strip() for v in vides_raw.split("\n") if v.strip()] if vides_raw else []

        entries.append(SubjectEntry(
            subject=assunto,
            sub_subject=sub_assunto,
            refs=refs,
            vides=vides,
        ))

    return SubjectIndex(entries=entries)


def _parse_dispositivos(raw: str, known_lettered: set[str] | None = None) -> list[SubjectRef]:
    """Converte string de dispositivos em lista de SubjectRef.

    Formatos aceitos:
    - "211-275"       → range de artigos
    - "175,II"        → artigo + inciso
    - "176,§10"       → artigo + parágrafo
    - "176,PU"        → artigo + parágrafo único
    - "176"           → artigo só
    - "LO:23"         → artigo de outra lei (prefixo:artigo)
    - "LO:23,II"      → artigo de outra lei + detalhe
    - Múltiplos separados por \\n
    """
    refs: list[SubjectRef] = []
    lines = raw.split("\n")

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Detect law prefix: "LO:23" or "LO:23,II"
        law_prefix = ""
        law_m = re.match(r"^([A-Z]{2,})\s*:\s*(.+)$", line)
        if law_m:
            law_prefix = law_m.group(1)
            line = law_m.group(2).strip()

        # Extract hint from parentheses at end of line: "13,I,b(propor privativamente)"
        hint = ""
        hint_m = re.search(r"\(([^)]+)\)\s*$", line)
        if hint_m:
            hint = hint_m.group(1).strip()
            line = line[:hint_m.start()].strip()

        # Range: "211-275"
        range_m = re.match(r"^(\d+)\s*[-–—]\s*(\d+)$", line)
        if range_m:
            start = int(range_m.group(1))
            end = int(range_m.group(2))
            for n in range(start, end + 1):
                refs.append(SubjectRef(art=str(n), law_prefix=law_prefix, hint=hint))
                # Inclui artigos letrados (ex: "212-A") cujo número base está no range
                if known_lettered:
                    for lettered in sorted(known_lettered):
                        m = re.match(r"^(\d+)-[A-Za-z]", lettered)
                        if m and int(m.group(1)) == n:
                            refs.append(SubjectRef(art=lettered, law_prefix=law_prefix, hint=hint))
            continue

        # Single or with detail: "175,II" or "176,§10" or "176"
        parts = line.split(",", 1)
        art = parts[0].strip()

        if not re.match(r"^\d+[-A-Za-z]*$", art):
            # Not a valid article reference, skip
            continue

        if len(parts) == 1:
            refs.append(SubjectRef(art=art, law_prefix=law_prefix, hint=hint))
        else:
            detail_raw = parts[1].strip()
            detail = _normalize_detail(detail_raw)
            refs.append(SubjectRef(art=art, detail=detail, law_prefix=law_prefix, hint=hint))

    return refs


def _normalize_detail(raw: str) -> str:
    """Normaliza detalhe do dispositivo para exibição.

    "II" → "II"
    "§10" → "§ 10"
    "§1" → "§ 1º"
    "PU" → "§ú"
    "§ú" → "§ú"
    "p1" → "§ 1º"
    """
    raw = raw.strip()

    if raw.upper() == "PU" or raw == "§ú":
        return "§ú"

    # §N → § Nº
    m = re.match(r"^[§Ss]\s*(\d+)$", raw)
    if m:
        num = m.group(1)
        return f"§ {num}º"

    # pN → § Nº
    m = re.match(r"^p(\d+)$", raw, re.IGNORECASE)
    if m:
        num = m.group(1)
        return f"§ {num}º"

    # Roman numeral (inciso)
    if re.match(r"^[IVXLC]+$", raw):
        return raw

    # Alínea
    if re.match(r"^[a-z]\)$", raw):
        return raw

    return raw
